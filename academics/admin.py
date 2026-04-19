from unfold.admin import ModelAdmin, TabularInline
from django.contrib import admin
from .models import Course, Session, Material, Evaluation, Result, Tutor, TutorEvaluation, IslamiyyaRegistration, IslamiyyaCourse, UserResourceSubmission, CompetitionResult,TimetableEntry
from django.http import HttpResponse
from django.urls import path, reverse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .forms import ExcelUploadForm
import openpyxl
from openpyxl.styles import PatternFill, Font
from django.utils.html import format_html
from openpyxl.utils import get_column_letter
from datetime import datetime
from .views import upload_timetable_excel
from django.utils import timezone
from core.services.webhooks import send_webhook
from django.utils.html import format_html

# Register your models here.

# ResultInline must be defined before EvaluationAdmin
class ResultInline(TabularInline):
    model = Result
    extra = 1
    readonly_fields = ('student_name', 'marks_obtained', 'grade')  # optional

# Define the view first
def upload_excel_view(request, evaluation_id):
    evaluation = get_object_or_404(Evaluation, id=evaluation_id)
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                messages.error(request, "File is empty.")
                return redirect('admin:academics_evaluation_changelist')
            created_count = 0
            for row in rows[1:]:
                if not any(row):
                    continue
                # Expect: Student Name, Reg No, Marks, Grade, Remarks (optional)
                if len(row) < 4:
                    continue
                name = str(row[0]) if row[0] else ''
                reg_no = str(row[1]) if len(row) > 1 and row[1] else ''
                marks = row[2] if len(row) > 2 else 0
                grade = str(row[3]) if len(row) > 3 and row[3] else ''
                remarks = str(row[4]) if len(row) > 4 and row[4] else ''
                try:
                    marks = float(marks)
                except (ValueError, TypeError):
                    marks = 0.0
                Result.objects.create(
                    evaluation=evaluation,
                    student_name=name,
                    registration_number=reg_no,
                    marks_obtained=marks,
                    grade=grade,
                    remarks=remarks
                )
                created_count += 1
            messages.success(request, f"Successfully imported {created_count} results.")
            return redirect('admin:academics_evaluation_changelist')
    else:
        form = ExcelUploadForm()
    context = {
        'form': form,
        'evaluation': evaluation,
        'title': f"Upload results for {evaluation}",
    }
    return render(request, 'admin/academics/upload_excel.html', context)

@admin.register(Evaluation)
class EvaluationAdmin(ModelAdmin):
    list_display = ('title', 'course', 'date', 'total_marks', 'upload_excel_button')
    list_filter = ('course', 'date')
    search_fields = ('title',)
    inlines = [ResultInline]

    def upload_excel_button(self, obj):
        url = reverse('admin:academics_evaluation_upload_excel', args=[obj.id])
        return format_html('<a class="button" href="{}">Upload Excel</a>', url)
    upload_excel_button.short_description = 'Upload Results'
    upload_excel_button.allow_tags = True

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<int:evaluation_id>/upload-excel/',
                 self.admin_site.admin_view(upload_excel_view),
                 name='academics_evaluation_upload_excel'),
        ]
        return custom_urls + urls

class SessionInline(TabularInline):
    model = Session
    extra = 1

class MaterialInline(TabularInline):
    model = Material
    extra = 1

class EvaluationInline(TabularInline):
    model = Evaluation
    extra = 1

class ResultInline(TabularInline):
    model = Result
    extra = 1
    readonly_fields = ('student_name', 'marks_obtained', 'grade')

@admin.register(Course)
class CourseAdmin(ModelAdmin):
    list_display = ('name', 'course_type', 'is_active')
    list_filter = ('course_type', 'is_active')
    search_fields = ('name',)
    prepopulated_fields = {'slug': ('name',)}
    inlines = [SessionInline, MaterialInline, EvaluationInline]
    filter_horizontal = ('tutors',)  # or use a widget for ManyToMany

@admin.register(Session)
class SessionAdmin(ModelAdmin):
    list_display = ('course', 'date', 'start_time', 'end_time', 'location')
    list_filter = ('course', 'date')
    search_fields = ('course__name',)

@admin.register(Material)
class MaterialAdmin(ModelAdmin):
    list_display = ('title', 'course', 'uploaded_at')
    list_filter = ('course',)
    search_fields = ('title',)

@admin.register(Result)
class ResultAdmin(ModelAdmin):
    list_display = ('student_name', 'evaluation', 'marks_obtained', 'grade')
    list_filter = ('evaluation__course',)
    search_fields = ('student_name', 'student_email')
    actions = ['import_from_excel']  # We'll add this later
    
    
# tutors evaluation admin
@admin.register(Tutor)
class TutorAdmin(ModelAdmin):
    list_display = ('name', 'is_active')
    list_filter = ('is_active',)
    search_fields = ('name',)

@admin.register(TutorEvaluation)
class TutorEvaluationAdmin(ModelAdmin):
    list_display = ('tutor', 'course', 'rating', 'student_name', 'submitted_at')
    list_filter = ('course', 'tutor', 'rating')
    search_fields = ('student_name', 'comments')
    actions = ['export_to_excel']

    def export_to_excel(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=tutor_evaluations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Evaluations"

        headers = ['Course', 'Tutor', 'Student Name', 'Rating', 'Comments', 'Submitted At']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws[f'{col_letter}1'].font = openpyxl.styles.Font(bold=True)

        for row_num, obj in enumerate(queryset, 2):
            ws[f'A{row_num}'] = str(obj.course)
            ws[f'B{row_num}'] = obj.tutor.name
            ws[f'C{row_num}'] = obj.student_name
            ws[f'D{row_num}'] = obj.rating
            ws[f'E{row_num}'] = obj.comments
            ws[f'F{row_num}'] = obj.submitted_at.strftime("%Y-%m-%d %H:%M")

        wb.save(response)
        return response
    export_to_excel.short_description = "Export selected evaluations to Excel"

    
@admin.register(TimetableEntry)
class TimetableEntryAdmin(ModelAdmin):
    list_display = ('course_name', 'entry_type', 'level', 'day', 'time_start', 'time_end', 'venue', 'is_active')
    list_filter = ('entry_type', 'level', 'day', 'is_active')
    search_fields = ('course_name', 'venue')
    list_editable = ('is_active',)
    actions = ['export_selected']  # optional

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-excel/', self.admin_site.admin_view(upload_timetable_excel), name='academics_timetableentry_upload'),
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['upload_button'] = True
        return super().changelist_view(request, extra_context=extra_context)
  

@admin.register(IslamiyyaCourse)
class IslamiyyaCourseAdmin(ModelAdmin):
    list_display = ('name', 'is_active')
    list_editable = ('is_active',)
    search_fields = ('name',)


# -------------------- Excel Export for islamiyyah application --------------------
def export_islamiyya_registrations_to_excel(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=islamiyya_registrations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Islamiyya Registrations"

    gold_fill = PatternFill(start_color="C9A84C", end_color="C9A84C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ['Application ID', 'Name', 'Registration Number', 'Email', 'Phone', 'Department', 'Gender', 'Level', 'Courses', 'Other Course', 'Submitted At', 'Verified', 'Verified At', 'WhatsApp Link']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = gold_fill
        cell.font = header_font

    for row_num, obj in enumerate(queryset, 2):
        courses_str = ', '.join([c.name for c in obj.courses.all()])
        ws.cell(row=row_num, column=1, value=obj.application_id)
        ws.cell(row=row_num, column=2, value=obj.name)
        ws.cell(row=row_num, column=3, value=obj.registration_number)
        ws.cell(row=row_num, column=4, value=obj.email)
        ws.cell(row=row_num, column=5, value=obj.phone)
        ws.cell(row=row_num, column=6, value=obj.department or '')
        ws.cell(row=row_num, column=7, value=obj.get_gender_display() if obj.gender else '')
        ws.cell(row=row_num, column=8, value=obj.get_level_display())
        ws.cell(row=row_num, column=9, value=courses_str)
        ws.cell(row=row_num, column=10, value=obj.other_course or '')
        ws.cell(row=row_num, column=11, value=obj.submitted_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row_num, column=12, value='Yes' if obj.is_verified else 'No')
        ws.cell(row=row_num, column=13, value=obj.verified_at.strftime('%Y-%m-%d %H:%M') if obj.verified_at else '')
    wb.save(response)
    return response
export_islamiyya_registrations_to_excel.short_description = "Export selected to Excel"

# -------------------- Bulk Actions --------------------
def mark_verified(modeladmin, request, queryset):
    queryset.update(is_verified=True, verified_at=datetime.now())
mark_verified.short_description = "Mark selected as verified"

def mark_unverified(modeladmin, request, queryset):
    queryset.update(is_verified=False, verified_at=None)
mark_unverified.short_description = "Mark selected as unverified"

# -------------------- Admin Class --------------------
@admin.register(IslamiyyaRegistration)
class IslamiyyaRegistrationAdmin(ModelAdmin):
    list_display = ('application_id', 'name', 'email', 'level', 'is_verified', 'submitted_at')
    list_filter = ('is_verified', 'level')
    search_fields = ('name', 'email', 'application_id', 'registration_number')
    actions = [mark_verified, mark_unverified, export_islamiyya_registrations_to_excel]
    filter_horizontal = ('courses',)
    
@admin.register(UserResourceSubmission)
class UserResourceSubmissionAdmin(ModelAdmin):
    list_display   = ('title', 'submitted_by', 'email', 'status', 'submitted_at', 'download_link')
    list_filter    = ('status', 'submitted_at')
    search_fields  = ('title', 'description', 'submitted_by', 'email')
    readonly_fields = ('submitted_at', 'submitted_by', 'email', 'title', 'description', 'file', 'download_link_in_form')
    ordering       = ('-submitted_at',)
    actions        = ['approve_submissions', 'reject_submissions']

    def approve_submissions(self, request, queryset):
        approved = 0
        for obj in queryset:
            if obj.status != 'approved' and not obj.email_sent:
                obj.status = 'approved'
                obj.reviewed_at = timezone.now()
                obj.save()
                # Send webhook
                submitter_name = obj.submitted_by if obj.submitted_by else 'User'
                send_webhook('resource_approved', {
                    'emial': obj.email,
                    'recipients': [obj.email],
                    'name': submitter_name,
                    'title': obj.title
                })
                obj.email_sent = True
                obj.save(update_fields=['email_sent'])
                approved += 1
        self.message_user(request, f"{approved} submission(s) approved.")

    def reject_submissions(self, request, queryset):
        from django.utils import timezone
        queryset.update(status='rejected', reviewed_at=timezone.now())
        self.message_user(request, f"{queryset.count()} submission(s) rejected.")
    reject_submissions.short_description = "Reject selected submissions"

    def download_link(self, obj):
        """Link in list view"""
        if obj.file:
            # If using Cloudinary, get the URL
            file_url = obj.file.url if hasattr(obj.file, 'url') else obj.file
            return format_html('<a href="{}" target="_blank" class="button">📥 Download</a>', file_url)
        return "—"
    download_link.short_description = 'Download File'

    def download_link_in_form(self, obj):
        """Link in change form (read-only field)"""
        if obj.file:
            file_url = obj.file.url if hasattr(obj.file, 'url') else obj.file
            return format_html('<a href="{}" target="_blank" class="button" style="background:#c9a84c; color:#1a1a1a; padding:5px 10px; border-radius:30px;">📥 Download File</a>', file_url)
        return "—"
    download_link_in_form.short_description = 'Download File'

@admin.register(CompetitionResult)
class CompetitionResultAdmin(ModelAdmin):
    list_display = ('event_name', 'category', 'position', 'participant_name', 'department', 'points', 'order')
    list_filter = ('event_name', 'category', 'year')
    search_fields = ('participant_name', 'event_name')
    list_editable = ('order',)
    actions = ['export_to_excel']

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-excel/', self.admin_site.admin_view(self.upload_excel), name='academics_competitionresult_upload'),
        ]
        return custom_urls + urls

    def upload_excel(self, request):
        """Import Competition Results from an Excel file."""
        if request.method == 'POST':
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))

                if len(rows) < 2:
                    messages.error(request, "File is empty or has only headers.")
                    return redirect('admin:academics_competitionresult_changelist')

                headers = rows[0]  # expected: event_name, category, position, participant_name, department, points, year, order
                expected = ['event_name', 'category', 'position', 'participant_name', 'department', 'points', 'year', 'order']
                # optional: validate headers loosely

                created_count = 0
                errors = []

                for idx, row in enumerate(rows[1:], start=2):
                    if not any(row):
                        continue
                    # Map by index (simplest – assume order as above)
                    try:
                        event_name = str(row[0]).strip() if row[0] else ''
                        category = str(row[1]).strip() if row[1] else ''
                        position = str(row[2]).strip() if row[2] else ''
                        participant_name = str(row[3]).strip() if row[3] else ''
                        department = str(row[4]).strip() if row[4] else ''
                        points = row[5] if row[5] else None
                        year = str(row[6]).strip() if row[6] else ''
                        order = int(row[7]) if row[7] else 0
                    except Exception as e:
                        errors.append(f"Row {idx}: parsing error - {e}")
                        continue

                    if not event_name or not participant_name:
                        errors.append(f"Row {idx}: event_name and participant_name are required.")
                        continue

                    # Convert points to Decimal
                    try:
                        points = float(points) if points else None
                    except:
                        points = None

                    CompetitionResult.objects.create(
                        event_name=event_name,
                        category=category,
                        position=position,
                        participant_name=participant_name,
                        department=department,
                        points=points,
                        year=year,
                        order=order,
                        is_active=True
                    )
                    created_count += 1

                if created_count:
                    messages.success(request, f"Successfully imported {created_count} competition results.")
                if errors:
                    for err in errors[:5]:
                        messages.error(request, err)
                    if len(errors) > 5:
                        messages.error(request, f"... and {len(errors)-5} more errors.")
                return redirect('admin:academics_competitionresult_changelist')
        else:
            form = ExcelUploadForm()

        context = {
            'form': form,
            'title': 'Upload Competition Results Excel',
        }
        return render(request, 'admin/academics/upload_excel.html', context)

    def export_to_excel(self, request, queryset):
        """Export selected competition results to Excel."""
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=competition_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Competition Results"

        headers = ['Event Name', 'Category', 'Position', 'Participant Name', 'Department', 'Points', 'Year', 'Order']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws[f'{col_letter}1'].font = openpyxl.styles.Font(bold=True)

        for row_num, obj in enumerate(queryset, 2):
            ws[f'A{row_num}'] = obj.event_name
            ws[f'B{row_num}'] = obj.category or ''
            ws[f'C{row_num}'] = obj.position
            ws[f'D{row_num}'] = obj.participant_name
            ws[f'E{row_num}'] = obj.department or ''
            ws[f'F{row_num}'] = obj.points if obj.points is not None else ''
            ws[f'G{row_num}'] = obj.year or ''
            ws[f'H{row_num}'] = obj.order

        wb.save(response)
        return response
    export_to_excel.short_description = "Export selected competition results to Excel"
    
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['upload_button'] = True
        return super().changelist_view(request, extra_context=extra_context)