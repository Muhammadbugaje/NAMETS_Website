from django.contrib import admin
from .models import ContactPhone, Patron, ExecutiveYear, Executive, Developer, Question, Answer, AboutPage, Skill, TutorApplication, MembershipApplication
from utils.admin_helpers import cloudinary_thumbnail

from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
# Register your models here.
@admin.register(Skill)
class SkillAdmin(admin.ModelAdmin):
    list_display = ['name', 'is_active']
    list_editable = ['is_active']
    search_fields = ['name']

class ExecutiveInline(admin.TabularInline):
    model = Executive
    extra = 1

@admin.register(Patron)
class PatronAdmin(admin.ModelAdmin):
    list_display = ('image_thumbnail', 'name', 'designation', 'hierarchy_order', 'is_active')
    list_editable = ('hierarchy_order', 'is_active')
    search_fields = ('name', 'designation')

    def image_thumbnail(self, obj):
        return cloudinary_thumbnail(obj.image)
    image_thumbnail.short_description = 'Photo'

@admin.register(ExecutiveYear)
class ExecutiveYearAdmin(admin.ModelAdmin):
    list_display = ('image_thumbnail', 'year_label', 'display_order', 'is_active')
    list_editable = ('display_order', 'is_active')
    inlines = [ExecutiveInline]

    def image_thumbnail(self, obj):
        return cloudinary_thumbnail(obj.photo)
    image_thumbnail.short_description = 'Photo'    

@admin.register(Executive)
class ExecutiveAdmin(admin.ModelAdmin):
    list_display = ('image_thumbnail', 'name', 'role', 'year', 'display_order', 'is_active')
    list_filter = ('year', 'is_active')
    list_editable = ('display_order', 'is_active')
    search_fields = ('name', 'role')

    def image_thumbnail(self, obj):
        return cloudinary_thumbnail(obj.photo)
    image_thumbnail.short_description = 'Photo'

@admin.register(Developer)
class DeveloperAdmin(admin.ModelAdmin):
    list_display = ('image_thumbnail', 'name', 'role', 'display_order', 'is_active')
    list_editable = ('display_order', 'is_active')

    def image_thumbnail(self, obj):
        return cloudinary_thumbnail(obj.photo)
    image_thumbnail.short_description = 'Photo'

@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ('name', 'category', 'is_public', 'submitted_at')
    list_filter = ('is_public', 'category')
    search_fields = ('name', 'question_text')
    actions = ['mark_public']

    def mark_public(self, request, queryset):
        queryset.update(is_public=True)
    mark_public.short_description = "Mark selected questions as public"

@admin.register(Answer)
class AnswerAdmin(admin.ModelAdmin):
    list_display = ('question', 'responded_by', 'responded_at')
    search_fields = ('question__question_text',)

@admin.register(AboutPage)
class AboutPageAdmin(admin.ModelAdmin):
    # singleton – only one instance allowed
    def has_add_permission(self, request):
        if AboutPage.objects.exists():
            return False
        return super().has_add_permission(request)
    

def export_tutor_applications_to_excel(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=tutor_applications_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tutor Applications"

    headers = ['Name', 'Reg Number', 'Department', 'CGPA', 'Email', 'Phone', 'Campus Residence',
               'Strong Courses', 'Preferred Course', 'Comfort Level', 'Teaching Skill Rating',
               'Recommendations', 'Past Experience', 'Availability', 'Has Materials',
               'Collaborate on Materials', 'Submitted At', 'Processed']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = openpyxl.styles.Font(bold=True)

    for row_num, obj in enumerate(queryset, 2):
        ws[f'A{row_num}'] = obj.name
        ws[f'B{row_num}'] = obj.reg_number
        ws[f'C{row_num}'] = obj.department
        ws[f'D{row_num}'] = str(obj.cgpa) if obj.cgpa else ''
        ws[f'E{row_num}'] = obj.email
        ws[f'F{row_num}'] = obj.phone
        ws[f'G{row_num}'] = 'Yes' if obj.campus_residence else 'No'
        ws[f'H{row_num}'] = obj.strong_courses
        ws[f'I{row_num}'] = obj.preferred_course
        ws[f'J{row_num}'] = obj.comfort_level
        ws[f'K{row_num}'] = obj.teaching_skill_rating or ''
        ws[f'L{row_num}'] = obj.recommendations
        ws[f'M{row_num}'] = obj.past_experience
        ws[f'N{row_num}'] = obj.availability
        ws[f'O{row_num}'] = 'Yes' if obj.has_materials else 'No'
        ws[f'P{row_num}'] = 'Yes' if obj.collaborate_on_materials else 'No'
        ws[f'Q{row_num}'] = obj.submitted_at.strftime('%Y-%m-%d %H:%M')
        ws[f'R{row_num}'] = 'Yes' if obj.is_processed else 'No'

    wb.save(response)
    return response
export_tutor_applications_to_excel.short_description = "Export selected to Excel"

def export_membership_applications_to_excel(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=membership_applications_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Membership Applications"

    headers = ['Name', 'Email', 'Phone', 'Gender', 'Reg Number', 'Department', 'Campus Residence',
               'Islamic Knowledge', 'Quran Memorization', 'Skills', 'Other Skill',
               'How Work with People', 'Recommendations', 'Submitted At', 'Processed']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = openpyxl.styles.Font(bold=True)

    for row_num, obj in enumerate(queryset, 2):
        ws[f'A{row_num}'] = obj.name
        ws[f'B{row_num}'] = obj.email
        ws[f'C{row_num}'] = obj.phone
        ws[f'D{row_num}'] = obj.get_gender_display()
        ws[f'E{row_num}'] = obj.reg_number
        ws[f'F{row_num}'] = obj.department
        ws[f'G{row_num}'] = 'Yes' if obj.campus_residence else 'No'
        ws[f'H{row_num}'] = obj.get_islamic_knowledge_display()
        ws[f'I{row_num}'] = obj.get_quran_memorization_display()
        # skills as comma-separated list
        skills_str = ', '.join([skill.name for skill in obj.skills.all()])
        ws[f'J{row_num}'] = skills_str
        ws[f'K{row_num}'] = obj.other_skill
        ws[f'L{row_num}'] = obj.how_work_with_people
        ws[f'M{row_num}'] = obj.recommendations
        ws[f'N{row_num}'] = obj.submitted_at.strftime('%Y-%m-%d %H:%M')
        ws[f'O{row_num}'] = 'Yes' if obj.is_processed else 'No'

    wb.save(response)
    return response
export_membership_applications_to_excel.short_description = "Export selected to Excel"

class TutorApplicationAdmin(admin.ModelAdmin):
    list_display = ['name', 'reg_number', 'email', 'preferred_course', 'submitted_at', 'is_processed']
    list_filter = ['is_processed', 'campus_residence']
    search_fields = ['name', 'email', 'reg_number']
    actions = [export_tutor_applications_to_excel]

class MembershipApplicationAdmin(admin.ModelAdmin):
    list_display = ['name', 'email', 'gender', 'islamic_knowledge', 'submitted_at', 'is_processed']
    list_filter = ['is_processed', 'gender', 'campus_residence']
    search_fields = ['name', 'email']
    filter_horizontal = ['skills']
    actions = [export_membership_applications_to_excel]

admin.site.register(TutorApplication, TutorApplicationAdmin)
admin.site.register(MembershipApplication, MembershipApplicationAdmin)

@admin.register(ContactPhone)
class ContactPhoneAdmin(admin.ModelAdmin):
    list_display = ['label', 'phone_number', 'order', 'is_active']
    list_editable = ['order', 'is_active']